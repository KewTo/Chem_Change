import pyautogui as auto
import pygetwindow as gw
import time
import os
from glob import glob
from openpyxl import load_workbook
import xlwings as xw
import re

# Automated 9904 disposition list for COAT Module. This is intended to grab the excel file from the Intel website, apply
# certain filters and formula in conjuncture with previous 9904 dispo list from the past. With the previous dispo list
# and VLOOKUP, it will apply which BLANKs belong to which engineer. It will sort and color code names for easier view.


def grab_file():
    win = gw.getWindowsWithTitle('New Tab - Google Chrome')[0]
    win.activate()
    auto.hotkey('ctrl', 't')
    time.sleep(2)
    auto.typewrite('https://imosc-ebiz.intel.com/imobi/module_stores.asp')
    auto.press('enter')
    time.sleep(30)
    for index in range(15):
        auto.press('tab')
        time.sleep(0.1)
        print(index)
    for index in range(4):
        auto.press('down')
        time.sleep(0.1)
        print(index)
    time.sleep(2)
    auto.click()
    win1 = gw.getWindowsWithTitle('Search by Module stores - Google Chrome')[0]
    win1.activate()
    time.sleep(120)
    for index in range(2):
        auto.press('tab')
        time.sleep(0.1)
    auto.press('enter')


def recent_file():
    # Find latest Excel file
    filepath = r'C:\Users\kevinto\Downloads\Search by Module Stores*.xlsx'
    latest_file = max(glob(filepath), key=os.path.getmtime)
    return latest_file


wb = load_workbook(recent_file())
ws = wb.active
ws.delete_cols(10)
wb.save(recent_file())
time.sleep(2)
xw.Book(recent_file())
Dispo_Dimensions = ws.dimensions


def filter_headers():
    # Find all header titles column number for filtering
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'ENG_LOT_OWNER':
                ENG_LOT_OWNER1 = cell.coordinate
    return ENG_LOT_OWNER1


def headers(x: str):
    header = []
    for i in range(1, ws.max_column):
        header.append(ws.cell(row=2, column=i).value)
    index = header.index(x) + 1
    return index


def copy_info():
    # Copy previous 9904 dispo list to a new sheet
    info_workbook = sorted(glob(r'C:\Users\kevinto\Downloads\Search by Module Stores*.xlsx'), key=os.path.getmtime)[-2]
    wb1 = xw.Book(info_workbook)  # Initial file with data
    wb2 = xw.Book(recent_file())  # Target file
    ws1 = wb1.sheets[0]  # [1]
    ws1.api.Copy(After=wb2.sheets[0].api)
    wb2.sheets[1].name = 'Sheet2'

    # Change to Sheet1 and maximize excel window screen
    time.sleep(1)
    auto.press('F6')
    auto.press('left')
    auto.press('enter')
    time.sleep(1)
    auto.hotkey('win', 'up')
    time.sleep(0.5)


def vlookup():
    # Apply VLOOKUP Formula on Status Column
    for x in range(2, len(xw.sheets[0].range('K1' + ':K' + str(ws.max_row)).rows)):
        xw.sheets[0].range('K' + str(xw.sheets[0].range('K1' + ':K' + str(ws.max_row))[x].row)).value = '=VLOOKUP(C' + str(
            x + 1) + ', Sheet2' + '!' + 'A:F, 6, FALSE)'


def copy_paste():
    # Copy vlookup on Status to ENG_LOT_OWNER
    head = []
    for i in range(2, len(xw.sheets[0].range('K1' + ':K' + str(ws.max_row)).rows)):
        if xw.sheets[0].range('K1' + ':K' + str(ws.max_row))[i].value is not None:
            head.append('K' + str(xw.sheets[0].range('K1' + ':K' + str(ws.max_row))[i].row))
    for i in head:
        xw.sheets[0].range(i).value = xw.sheets[0].range(i).options(ndim=2).value
        xw.sheets[0].range(i).copy()
        xw.sheets[0].range('J' + str(i[1:])).paste()


def owner_name():
    # Color code ENG_LOT_OWNER Color = (R,G,B)
    for index, elem in enumerate(xw.sheets[0].range('J1' + ':J' + str(ws.max_row)).value, start=1):
        if re.match(r'^BWOLSON',  str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (204, 0, 0)
        elif re.match(r'^GLUU', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (204, 102, 102)
        elif re.match(r'^HJAVAID', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (204, 0, 204)
        elif re.match(r'^IECHOLS', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (102, 204, 204)
        elif re.match(r'^JABELARD', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (204, 0, 102)
        elif re.match(r'^JKBOSWOR', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (204, 204, 0)
        elif re.match(r'^JRNISKAL', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (0, 204, 0)
        elif re.match(r'^PSRIVAS2', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (0, 204, 204)
        elif re.match(r'^SCPRICE', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (0, 0, 204)
        elif re.match(r'^YUNPINGF', str(elem), flags=re.IGNORECASE):
            xw.Range(filter_headers()[0] + str(index)).color = (102, 0, 204)
        elif elem == 'ENG_LOT_OWNER':
            pass
        else:
            xw.Range(filter_headers()[0] + str(index)).color = (204, 204, 102)


def delete_extra():
    # Delete more GOLDEN_MASK without Y
    for index, elem in reversed(list(enumerate(xw.sheets[0].range('C1' + ':C' + str(ws.max_row)).value, start=1))):
        if elem == "BLNK408215": #PEB GOLDEN MASK MONITOR
            xw.sheets[0].range('A' + str(index) + ':' + 'V' + str(index)).delete()
        elif re.match(r'.*CUPWASH', str(elem)):
            xw.sheets[0].range('A' + str(index) + ':' + 'V' + str(index)).delete()
        elif re.match(r'.*WARM1', str(elem)):
            xw.sheets[0].range('A' + str(index) + ':' + 'V' + str(index)).delete()
    # Delete unnecessary rows/columns
    xw.sheets[0].range('A1:V1').delete()
    xw.sheets[0].range('I1:I' + str(xw.sheets[0].range(Dispo_Dimensions).current_region.last_cell.row)).delete()
    xw.sheets[0].range('E1:E' + str(xw.sheets[0].range(Dispo_Dimensions).current_region.last_cell.row)).delete()
    xw.sheets[0].range('B1:B' + str(xw.sheets[0].range(Dispo_Dimensions).current_region.last_cell.row)).delete()
    xw.sheets[0].range('A1:A' + str(xw.sheets[0].range(Dispo_Dimensions).current_region.last_cell.row)).delete()


def stall_out():
    import time
    start = time.time()
    input("Press any key to continue:")
    end = time.time()
    elapsed = end-start
    elapsed = elapsed / 60
    print(str(elapsed) + " minutes has elapsed")


grab_file()
copy_info()
vlookup()
copy_paste()
stall_out()
owner_name()
delete_extra()

if __name__ == '__main__':
    pass
